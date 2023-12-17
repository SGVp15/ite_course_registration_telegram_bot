import openpyxl
import xlrd


# Open the output text file
def xlsx_to_txt(input_file, output_file):
    with open(output_file, 'w') as f:
        # Load the Excel file
        wb = openpyxl.load_workbook(input_file)

        # Select the worksheet you want to convert
        sheet_names = wb.sheetnames
        ws = sheet_names[0]

        # Loop through each row in the worksheet
        for row in ws.iter_rows():
            # Join the values in each cell with a tab character
            row_values = [str(cell.value) for cell in row]
            row_text = '\t'.join(row_values)

            # Write the row to the text file
            f.write(row_text + '\n')


def read_xlsx(input_file) -> str:
    wb = openpyxl.load_workbook(input_file)

    # Select the worksheet you want to convert
    out = ''
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Loop through each row in the worksheet
        for rows in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row):
            s = ''
            # Join the values in each cell with a tab character
            rows = [v for v in rows if v.value]
            for row in rows:
                s += str(row.value) + '\t'
            out += s + '\n'
    return out


def read_xls(input_file) -> str:
    wb = xlrd.open_workbook(input_file)
    out = ''
    for sheet_name in wb.sheet_names():
        # Select the worksheet by name
        ws = wb.sheet_by_name(sheet_name)

        s = ''
        # Loop through each row and column, and print the cell value
        for row in range(ws.nrows):
            row_value = ''
            for col in range(ws.ncols):
                cell_value = ws.cell_value(row, col)
                row_value += f'{cell_value}\t'
            s += f"{row_value}\n"
        out += s
    return out
